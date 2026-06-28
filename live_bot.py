import time
import datetime
import os
import winsound
from src.data_fetcher import DataFetcher
from src.smc_analyzer import SMCAnalyzer
from src.utils import logger
from src.config import (
    CANDLE_FETCH_COUNT,
    CONFLUENCE_THRESHOLD,
    MARKET_OPEN_TIME,
    MARKET_CLOSE_TIME,
)

def format_upstox_time(iso_time_str):
    """Converts Upstox ISO timestamp to a readable format like 22-Jun-2026 11:33 AM."""
    try:
        dt = datetime.datetime.fromisoformat(iso_time_str)
        return dt.strftime('%d-%b-%Y %I:%M %p')
    except Exception:
        return iso_time_str

def update_excel_report(trades, filename="Live_Sniper_Tracker.xlsx"):
    """Generates and updates the live color-coded Excel tracker, preserving past days on separate tabs."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        
        # Get today's date to use as the Sheet name
        today_str = datetime.datetime.now().strftime('%d-%b-%Y')
        
        # Load existing workbook if it exists, otherwise create a new one
        if os.path.exists(filename):
            wb = openpyxl.load_workbook(filename)
        else:
            wb = openpyxl.Workbook()
            # Remove the default empty sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        
        # If today's sheet already exists, we select it and clear it to refresh the live statuses.
        # If it doesn't exist, we create a fresh tab for today.
        if today_str in wb.sheetnames:
            ws = wb[today_str]
            ws.delete_rows(1, ws.max_row) 
        else:
            ws = wb.create_sheet(title=today_str)
        
        headers = ['Instrument', 'Detected Time', 'Candle Time', 'Type', 'Entry Zone', 'Avg Entry', 'Stop Loss', 'Take Profit', 'Score', 'Status']
        ws.append(headers)
        
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(color="006100", bold=True)
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006", bold=True)
        
        for row_idx, t in enumerate(trades, start=2):
            ws.cell(row=row_idx, column=1, value=t['instrument'])
            ws.cell(row=row_idx, column=2, value=t['detected_time'])
            ws.cell(row=row_idx, column=3, value=t['candle_time'])
            ws.cell(row=row_idx, column=4, value=t['type'])
            ws.cell(row=row_idx, column=5, value=f"{t['bottom']} - {t['top']}")
            ws.cell(row=row_idx, column=6, value=t['entry_price'])
            ws.cell(row=row_idx, column=7, value=t['stop_loss'])
            ws.cell(row=row_idx, column=8, value=t['take_profit'])
            ws.cell(row=row_idx, column=9, value=t['score'])
            
            status_cell = ws.cell(row=row_idx, column=10, value=t['status'])
            
            for col in range(1, 11):
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")
                
            if "SUCCESS" in t['status']:
                status_cell.fill = green_fill
                status_cell.font = green_font
            elif "FAILED" in t['status']:
                status_cell.fill = red_fill
                status_cell.font = red_font
                
        # Auto-adjust widths to prevent ###### display bugs
        for col_letter in ['B', 'C', 'E', 'J']:
            ws.column_dimensions[col_letter].width = 22
        for col_letter in ['A', 'D']:
            ws.column_dimensions[col_letter].width = 15
            
        wb.save(filename)
    except Exception as e:
        logger.error(f"Failed to save Excel: {e}")

def main():
    data_fetcher = DataFetcher()
    target_indices = ["NIFTY", "BANKNIFTY", "NIFTYIT", "SENSEX"]
    last_alerted_fvg_time = {index: None for index in target_indices}
    
    # Active Memory for tracking open trades
    active_trades = [] 
    
    print("\n" + "="*60)
    print("🎯 Multi-Index SMC Sniper Bot Active")
    print(f"Tracking: {', '.join(target_indices)}")
    print(f"Threshold: Score >= {CONFLUENCE_THRESHOLD}")
    print("Listening for High-Confluence Setups...")
    print("="*60 + "\n")

    while True:
        now = datetime.datetime.now()
        current_time_str = now.strftime('%Y-%m-%d %H:%M:%S')
        
        if now.weekday() >= 5:
            print(f"[{current_time_str}] Weekend. Market is closed. Sleeping for 1 hour...")
            time.sleep(3600)
            continue
            
        market_start = now.replace(hour=MARKET_OPEN_TIME.hour, minute=MARKET_OPEN_TIME.minute, second=0, microsecond=0)
        market_end = now.replace(hour=MARKET_CLOSE_TIME.hour, minute=MARKET_CLOSE_TIME.minute, second=0, microsecond=0)
        
        if now < market_start or now > market_end:
            print(f"[{current_time_str}] Outside market hours. Sleeping for 5 minutes...")
            time.sleep(300)
            continue

        for index_name in target_indices:
            print(f"[{current_time_str}] Scanning {index_name}...")
            symbol = data_fetcher.get_current_futures_symbol(index_name)
            
            if symbol:
                candles = data_fetcher.get_recent_candles(symbol, interval="1minute", count=CANDLE_FETCH_COUNT)
                
                if candles:
                    latest_candle = candles[-1]
                    
                    # --- LIVE TP/SL MONITORING ---
                    status_changed = False
                    for t in active_trades:
                        if t['status'] == 'PENDING' and t['instrument'] == index_name:
                            if t['type'] == 'BULLISH_FVG':
                                if latest_candle['low'] <= t['stop_loss']:
                                    t['status'] = 'FAILED (SL Hit)'
                                    status_changed = True
                                elif latest_candle['high'] >= t['take_profit']:
                                    t['status'] = 'SUCCESS (TP Hit)'
                                    status_changed = True
                            elif t['type'] == 'BEARISH_FVG':
                                if latest_candle['high'] >= t['stop_loss']:
                                    t['status'] = 'FAILED (SL Hit)'
                                    status_changed = True
                                elif latest_candle['low'] <= t['take_profit']:
                                    t['status'] = 'SUCCESS (TP Hit)'
                                    status_changed = True
                    
                    if status_changed:
                        update_excel_report(active_trades)

                    # --- NEW SETUP SCANNING ---
                    fvgs = SMCAnalyzer.analyze_fvgs(candles)
                    enriched_fvgs = SMCAnalyzer.enrich_with_confluence(candles, fvgs)
                    sniper_setups = [f for f in enriched_fvgs if f['confluence_count'] >= CONFLUENCE_THRESHOLD]
                    
                    if sniper_setups:
                        latest_setup = sniper_setups[-1] 
                        
                        if latest_setup['time'] != last_alerted_fvg_time[index_name]:
                            last_alerted_fvg_time[index_name] = latest_setup['time']
                            
                            c2_idx = latest_setup['creation_index'] - 1
                            if c2_idx < len(candles):
                                displacement_candle = candles[c2_idx]
                                vol_sma = max(displacement_candle.get('vol_sma_20', 1), 1) 
                                vol_surge = round(displacement_candle.get('volume', 0) / vol_sma, 1)
                            else:
                                vol_surge = 0.0

                            formatted_candle_time = format_upstox_time(latest_setup['time'])
                            
                            # Calculate SL & TP based on C1
                            c1_idx = latest_setup['creation_index'] - 2
                            entry_price = round((latest_setup['top'] + latest_setup['bottom']) / 2, 2)
                            
                            if c1_idx >= 0 and c1_idx < len(candles):
                                c1 = candles[c1_idx]
                                if latest_setup['type'] == 'BULLISH_FVG':
                                    stop_loss = float(c1['low'])
                                    risk = entry_price - stop_loss
                                    take_profit = round(entry_price + (risk * 2), 2)
                                else:
                                    stop_loss = float(c1['high'])
                                    risk = stop_loss - entry_price
                                    take_profit = round(entry_price - (risk * 2), 2)
                            else:
                                stop_loss, take_profit = 0.0, 0.0

                            try:
                                winsound.Beep(1000, 1500) 
                            except Exception:
                                pass 
                                
                            passed_checks = []
                            if latest_setup['confluence']['htf_aligned']: passed_checks.append("HTF Bias")
                            if latest_setup['confluence']['zone_aligned']: passed_checks.append("PD Zone")
                            if latest_setup['confluence']['liquidity_sweep']: passed_checks.append("Sweep")
                            if latest_setup['confluence']['ob_confluence']: passed_checks.append("Order Block")
                            passed_str = ", ".join(passed_checks) if passed_checks else "None"

                            print("\n" + "🔥"*20)
                            print(f"🚨 {index_name} SNIPER SETUP DETECTED! 🚨")
                            print(f"Detected At: {current_time_str}")
                            print(f"Type: {latest_setup['type']}")
                            print(f"Candle Time: {formatted_candle_time}")
                            print(f"Entry Zone: {latest_setup['bottom']} - {latest_setup['top']} (Avg: {entry_price})")
                            print(f"🛑 Stop Loss: {stop_loss}")
                            print(f"🎯 Take Profit (1:2): {take_profit}")
                            print(f"Volume Surge: {vol_surge}x")
                            print(f"Confluence Score: {latest_setup['confluence_count']}/4")
                            print(f"Passed Checks: [{passed_str}]")
                            print("🔥"*20 + "\n")
                            
                            # Save to memory and update Excel tracker
                            new_trade = {
                                'instrument': index_name,
                                'detected_time': current_time_str,
                                'candle_time': formatted_candle_time,
                                'type': latest_setup['type'],
                                'bottom': latest_setup['bottom'],
                                'top': latest_setup['top'],
                                'entry_price': entry_price,
                                'stop_loss': stop_loss,
                                'take_profit': take_profit,
                                'score': f"{latest_setup['confluence_count']}/4",
                                'status': 'PENDING'
                            }
                            active_trades.append(new_trade)
                            update_excel_report(active_trades)
                            print(f"💾 Trade saved to Tracker. Monitoring Live SL/TP...")
                            
            time.sleep(2)
        print("-" * 40)
        time.sleep(60)

if __name__ == "__main__":
    main()