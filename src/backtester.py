import sys
import os

# Add the project root to the Python path so absolute imports like 'from src...' work from anywhere
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import datetime
import time
import pandas as pd
from src.data_fetcher import DataFetcher
from src.smc_analyzer import SMCAnalyzer
from src.utils import logger
from src.config import NIFTY_SYMBOL, BANKNIFTY_SYMBOL

class SMCBacktester:
    def __init__(self):
        self.data_fetcher = DataFetcher()
        self.analyzer = SMCAnalyzer()
        
    def fetch_historical_intraday(self, symbol, interval="1minute", days_back=5):
        """Fetches historical intraday candles for backtesting."""
        logger.info(f"Fetching last {days_back} days of {interval} data for {symbol}...")
        
        end_date = datetime.date.today()
        start_date = end_date - datetime.timedelta(days=days_back)
        
        # Upstox API format: YYYY-MM-DD
        to_date = end_date.strftime('%Y-%m-%d')
        from_date = start_date.strftime('%Y-%m-%d')
        
        response = self.data_fetcher.upstox_handler.get_historical_data(
            symbol=symbol, 
            interval=interval, 
            from_date=from_date, 
            to_date=to_date
        )
        
        if not response or response.get('status') != 'success':
            logger.error("Failed to fetch historical backtest data.")
            return []
            
        candles_data = response.get('data', {}).get('candles', [])
        
        formatted_candles = []
        for c in candles_data:
            formatted_candles.append({
                'time': c[0],
                'open': float(c[1]),
                'high': float(c[2]),
                'low': float(c[3]),
                'close': float(c[4]),
                'volume': float(c[5])
            })
            
        # SMART SORT: Ensure chronological order (oldest to newest)
        if formatted_candles and formatted_candles[0]['time'] > formatted_candles[-1]['time']:
            formatted_candles.reverse()
            
        if formatted_candles:
            # Calculate intraday VWAP per date and 20-period volume SMA for futures
            df = pd.DataFrame(formatted_candles)
            df['datetime'] = pd.to_datetime(df['time'])
            df['date'] = df['datetime'].dt.date
            df['typical_price'] = (df['high'] + df['low'] + df['close']) / 3
            df['tp_vol'] = df['typical_price'] * df['volume']
            df['cum_tp_vol'] = df.groupby('date')['tp_vol'].cumsum()
            df['cum_vol'] = df.groupby('date')['volume'].cumsum()
            df['vwap'] = df['cum_tp_vol'] / df['cum_vol']
            df['vol_sma_20'] = df['volume'].rolling(window=20).mean()

            for i, row in df.iterrows():
                formatted_candles[i]['vwap'] = row['vwap']
                formatted_candles[i]['vol_sma_20'] = row['vol_sma_20']

            logger.info(f"✓ Downloaded {len(formatted_candles)} historical candles.")
            logger.info(f"Data Range: {formatted_candles[0]['time']} to {formatted_candles[-1]['time']}")
            
        return formatted_candles

    def _format_readable_time(self, raw_time):
        """Converts Upstox API time to a human-readable 12-hour AM/PM format."""
        if "UNMITIGATED" in str(raw_time):
            return raw_time
        try:
            dt = datetime.datetime.fromisoformat(str(raw_time))
            return dt.strftime('%d-%b-%Y %I:%M %p')
        except Exception:
            return raw_time

    def _is_in_killzone(self, raw_time):
        """Checks if the FVG was formed during high-probability institutional hours."""
        try:
            dt = datetime.datetime.fromisoformat(raw_time)
            time_val = dt.time()
            morning_start = datetime.time(9, 15)
            morning_end = datetime.time(11, 0)
            afternoon_start = datetime.time(13, 0)
            afternoon_end = datetime.time(15, 0)
            
            return (morning_start <= time_val <= morning_end) or (afternoon_start <= time_val <= afternoon_end)
        except Exception:
            return True # If parsing fails, don't filter it out

    def run_fvg_backtest(self, candles):
        """Scans for Highly Filtered Sniper FVGs and tracks mitigation behavior."""
        logger.info("Running Ultra-Filtered Sniper Backtest...")
        
        all_fvgs = self._local_analyze_fvgs(candles)
        logger.info(f"Total SNIPER FVGs identified (Vol Surge & Killzone Filtered): {len(all_fvgs)}")
        
        results = []
        
        for fvg in all_fvgs:
            fvg_time = fvg['time']
            top = fvg['top']
            bottom = fvg['bottom']
            fvg_type = fvg['type']
            gap_size = fvg['gap_size']
            vol_multiplier = fvg.get('vol_multiplier', 1)
            
            creation_index = next((i for i, c in enumerate(candles) if c['time'] == fvg_time), -1)
            
            mitigated_time = "UNMITIGATED (Still Open!)"
            status = "PENDING (Open)"
            
            if creation_index != -1:
                for future_candle in candles[creation_index + 2:]:
                    # Candle Anatomy Math
                    total_range = future_candle['high'] - future_candle['low']
                    if total_range == 0: continue
                    
                    upper_wick = future_candle['high'] - max(future_candle['open'], future_candle['close'])
                    lower_wick = min(future_candle['open'], future_candle['close']) - future_candle['low']
                    
                    upper_wick_percent = upper_wick / total_range
                    lower_wick_percent = lower_wick / total_range

                    # 🟢 BULLISH FVG Mitigation Rules
                    if fvg_type == 'BULLISH_FVG':
                        if future_candle['low'] <= top: # Price entered the Sniper Zone
                            if future_candle['close'] < bottom:
                                status = "FAILED (Invalidated)"
                                mitigated_time = future_candle['time']
                                break
                            if lower_wick_percent >= 0.4:
                                status = "SNIPER SUCCESS"
                                mitigated_time = future_candle['time']
                                break

                    # 🔴 BEARISH FVG Mitigation Rules
                    elif fvg_type == 'BEARISH_FVG':
                        if future_candle['high'] >= bottom: # Price entered the Sniper Zone
                            if future_candle['close'] > top:
                                status = "FAILED (Invalidated)"
                                mitigated_time = future_candle['time']
                                break
                            if upper_wick_percent >= 0.4:
                                status = "SNIPER SUCCESS"
                                mitigated_time = future_candle['time']
                                break
                        
            results.append({
                "Type": fvg_type,
                "Created At": self._format_readable_time(fvg_time),
                "FVG Top": top,
                "FVG Bottom": bottom,
                "Gap Size": gap_size,
                "Vol Surge": f"{vol_multiplier}x",
                "Status": status,
                "Mitigated At": self._format_readable_time(mitigated_time)
            })
            
        return results
        
    def _local_analyze_fvgs(self, candles, size_filter=10.0, vol_filter=1.5):
        """Advanced FVG detection using VWAP, Size, Volume Surge, and Time Filters."""
        fvgs = []
        if len(candles) < 50:
            return fvgs

        for i in range(50, len(candles) - 2):
            c1 = candles[i]
            c2 = candles[i+1]  # Displacement Candle
            c3 = candles[i+2]

            # VWAP and volume SMA based filters
            vwap = c2.get('vwap', c2['close'])
            vol_sma = c2.get('vol_sma_20', 1)
            if pd.isna(vol_sma) or vol_sma == 0:
                vol_sma = 1
            vol_multiplier = round(c2['volume'] / vol_sma, 1)

            # PRE-FILTER: Ignore if not in the institutional killzone
            if not self._is_in_killzone(c2['time']):
                continue

            # 🟢 Bullish FVG
            if c3['low'] > c1['high'] and c2['close'] > c2['open']:
                gap_size = c3['low'] - c1['high']
                # Requires: Massive Gap + Price above VWAP + Volume Surge
                if gap_size >= size_filter and c2['close'] > vwap and vol_multiplier >= vol_filter:
                    fvgs.append({
                        'type': 'BULLISH_FVG',
                        'time': c2['time'],
                        'top': c3['low'],
                        'bottom': c1['high'],
                        'gap_size': round(gap_size, 2),
                        'vol_multiplier': vol_multiplier
                    })

            # 🔴 Bearish FVG
            elif c3['high'] < c1['low'] and c2['close'] < c2['open']:
                gap_size = c1['low'] - c3['high']
                # Requires: Massive Gap + Price below VWAP + Volume Surge
                if gap_size >= size_filter and c2['close'] < vwap and vol_multiplier >= vol_filter:
                    fvgs.append({
                        'type': 'BEARISH_FVG',
                        'time': c2['time'],
                        'top': c1['low'],
                        'bottom': c3['high'],
                        'gap_size': round(gap_size, 2),
                        'vol_multiplier': vol_multiplier
                    })
        return fvgs

    def print_report(self, results):
        """Generates a color-coded Excel spreadsheet and prints a console summary."""
        df = pd.DataFrame(results)
        
        print("\n" + "="*95)
        print("    SMC FUTURES REPORT (VWAP + VOLUME SURGE + KILLZONES + WICKS)")
        print("="*95)
        if df.empty:
            print("No FVGs passed the strict Institutional Filters!")
        else:
            print(df.to_string(index=False))
        print("="*95)
        
        export_filename = "SMC_Ultra_Sniper_Report.xlsx"
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sniper Entries"

            headers = ["Trade Type", "Created Time", "FVG Top", "FVG Bottom", "Gap Size", "Vol Surge", "Result Status", "Trigger Time"]
            ws.append(headers)
            
            header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            green_font = Font(color="006100", bold=True)
            
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            red_font = Font(color="9C0006", bold=True)
            
            for row_idx, r in enumerate(results, start=2):
                ws.cell(row=row_idx, column=1, value=r["Type"])
                ws.cell(row=row_idx, column=2, value=r["Created At"])
                ws.cell(row=row_idx, column=3, value=r["FVG Top"])
                ws.cell(row=row_idx, column=4, value=r["FVG Bottom"])
                ws.cell(row=row_idx, column=5, value=r["Gap Size"])
                ws.cell(row=row_idx, column=6, value=r.get("Vol Surge", ""))
                
                status_cell = ws.cell(row=row_idx, column=7, value=r["Status"])
                ws.cell(row=row_idx, column=8, value=r["Mitigated At"])
                
                for col in range(1, 9):
                    ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")
                
                if "SUCCESS" in r["Status"]:
                    status_cell.fill = green_fill
                    status_cell.font = green_font
                elif "FAILED" in r["Status"]:
                    status_cell.fill = red_fill
                    status_cell.font = red_font

            for col_letter in ['A', 'B', 'H']:
                ws.column_dimensions[col_letter].width = 24
            for col_letter in ['C', 'D', 'E', 'F']:
                ws.column_dimensions[col_letter].width = 14
            ws.column_dimensions['G'].width = 22

            wb.save(export_filename)
            print(f"\n✅ ULTRA-ADVANCED EXCEL SHEET GENERATED: '{export_filename}'")

        except ImportError:
            df.to_csv("SMC_Ultra_Sniper_Report.csv", index=False)
            print("\n⚠️ Note: 'openpyxl' is not installed. Saved as CSV instead.")

if __name__ == "__main__":
    backtester = SMCBacktester()

    # Automatically fetch the active Bank Nifty Future instrument key from Upstox
    target_symbol = backtester.data_fetcher.get_current_futures_symbol("BANKNIFTY")

    if target_symbol:
        historical_candles = backtester.fetch_historical_intraday(target_symbol, interval="1minute", days_back=5)
        if historical_candles:
            report = backtester.run_fvg_backtest(historical_candles)
            backtester.print_report(report)
    else:
        print("❌ Failed to fetch the active Futures symbol. Cannot run backtest.")